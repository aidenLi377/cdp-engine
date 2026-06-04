import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


class TestExcelBuilder(unittest.TestCase):
    def setUp(self):
        from cdp_backend.rpa_agent.excel_builder import ExcelBuilder
        self.builder = ExcelBuilder()
        self.sample_rows = [
            {
                "大类": "用户特征",
                "标签类型": "基础特征",
                "标签名称": "居住城市",
                "特征明细": "杭州",
                "人群占比": "12.50%",
                "Rebase": "12.50%",
                "点击TGI": 1.15,
                "转化TGI": 0.98,
            },
            {
                "大类": "用户特征",
                "标签类型": "基础特征",
                "标签名称": "用户年龄",
                "特征明细": "25-30",
                "人群占比": "18.20%",
                "Rebase": "18.20%",
                "点击TGI": 1.42,
                "转化TGI": 1.21,
            },
            {
                "大类": "品类特征",
                "标签类型": "策略人群",
                "标签名称": "大快消策略人群",
                "特征明细": "美妆达人",
                "人群占比": "25.00%",
                "Rebase": "25.00%",
                "点击TGI": 2.10,
                "转化TGI": 1.85,
            },
        ]

    def test_build_excel_creates_valid_xlsx(self):
        rows = self.sample_rows
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "test.xlsx"
            self.builder.build(rows, str(path))
            self.assertTrue(path.exists())
            wb = load_workbook(str(path))
            ws = wb.active
            self.assertEqual(ws.title, "画像透视数据")
            self.assertEqual(ws.cell(1, 1).value, "大类")
            self.assertEqual(ws.cell(1, 5).value, "人群占比")
            self.assertEqual(ws.cell(2, 1).value, "用户特征")
            self.assertEqual(ws.cell(2, 3).value, "居住城市")
            self.assertEqual(ws.cell(2, 4).value, "杭州")

    def test_build_excel_header_frozen(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "test_frozen.xlsx"
            self.builder.build(self.sample_rows, str(path))
            wb = load_workbook(str(path))
            ws = wb.active
            self.assertEqual(ws.freeze_panes, "A2")

    def test_build_empty_rows_raises_value_error(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "empty.xlsx"
            with self.assertRaises(ValueError):
                self.builder.build([], str(path))


if __name__ == "__main__":
    unittest.main()
