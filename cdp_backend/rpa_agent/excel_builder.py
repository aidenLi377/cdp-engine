"""Generate styled .xlsx from DMP portrait data."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CATEGORY_FILLS = {
    "用户特征": PatternFill(start_color="E6F4FF", end_color="E6F4FF", fill_type="solid"),
    "品类特征": PatternFill(start_color="FFF2E8", end_color="FFF2E8", fill_type="solid"),
    "渠道特征": PatternFill(start_color="F6FFED", end_color="F6FFED", fill_type="solid"),
    "私域特征": PatternFill(start_color="F9F0FF", end_color="F9F0FF", fill_type="solid"),
}

HEADER_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="333333")
DATA_FONT = Font(name="Microsoft YaHei", size=10, color="555555")
THIN_BORDER = Border(
    left=Side(style="thin", color="EAEAEA"),
    right=Side(style="thin", color="EAEAEA"),
    top=Side(style="thin", color="EAEAEA"),
    bottom=Side(style="thin", color="EAEAEA"),
)
HEADER_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

COLUMNS = ["大类", "标签类型", "标签名称", "特征明细", "人群占比", "Rebase", "点击TGI", "转化TGI"]
COLUMN_WIDTHS = [14, 14, 20, 22, 12, 12, 10, 10]


class ExcelBuilder:
    def build(self, rows: list[dict], output_path: str) -> None:
        if not rows:
            raise ValueError("No data rows to export")

        wb = Workbook()
        ws = wb.active
        ws.title = "画像透视数据"

        for col_idx, (header, width) in enumerate(zip(COLUMNS, COLUMN_WIDTHS), 1):
            cell = ws.cell(1, col_idx, header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = HEADER_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

        for row_idx, row_data in enumerate(rows, 2):
            cat_fill = CATEGORY_FILLS.get(row_data.get("大类", ""))
            for col_idx, key in enumerate(COLUMNS, 1):
                value = row_data.get(key, "")
                cell = ws.cell(row_idx, col_idx, value)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")
                if col_idx == 1 and cat_fill:
                    cell.fill = cat_fill

        ws.row_dimensions[1].height = 28
        for row_idx in range(2, len(rows) + 2):
            ws.row_dimensions[row_idx].height = 22

        wb.save(output_path)
